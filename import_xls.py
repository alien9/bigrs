from openpyxl import load_workbook
import sys,os,re

def integer(v):
    try:
        i=int(v)
    except:
        return 'NULL'
    return i
def reprr(t):
    if t is None:
        return 'NULL'
    else:
        return repr(str(t))

path="/home/tiago/Dropbox/Dados e Cartografias - MobiLab/Mapas/Auxiliares/SAT dados Hori 7abril2016/"
fs=os.listdir(path)
fs=[l for l in fs if re.search("xlsx$",l)]
for f in fs:
    wb2 = load_workbook("%s%s"%(path,f,),guess_types=True)
    sheet=wb2.active
    if str(sheet.title)=="Veículos_":
        for row in sheet.rows:
            try:
                idade = int(row[5].value)
            except:
                idade='NULL'

            q="INSERT INTO public.veiculos(" \
              "id_acidente, id_veiculo, tipo_veiculo, placa, sexo_condutor,idade_condutor, categoria_habilitacao, usa_cinto_seguranca, estado_alcoolizacao,escolaridade" \
              ") VALUES (" \
              "'%s','%s','%s','%s','%s',%s,'%s','%s','%s','%s');"%(
                row[0].value,row[1].value,row[2].value,row[3].value,row[4].value,idade,row[6].value,row[7].value,row[8].value,row[9].value,
            )
            print(q)
    if str(sheet.title)=="Acidentes_":
        for row in sheet.rows:
            dias=str(row[5].value).split("/")
            if len(dias)==3:
                data="%s-%s-%s"%(dias[2],dias[1],dias[0],)
                dataehora="%s %s"%(data,str(row[6].value),)
            else:
                data=""
                dataehora=""
            try:
                q="INSERT INTO public.acidentes(" \
                  "id_acidente,cadloga,cadlogb,alt_num,referencia,data,hora,cod_acid,tipo_acidente,talao,sentido,automovel,moto,onibus,caminhao,bicicleta,moto_taxi,onibus_fretado,onibus_urbano,microonibus,van,vuc,caminhonete,carreta,jipe,carroca,outros,sem_informacao,feridos,mortos,dp,pista,det,distrito,fonte,princ_a,princ_b)" \
                  "VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);"%(
                    reprr(row[0].value),
                    reprr(row[1].value),
                    reprr(row[2].value),
                    reprr(row[3].value),
                    reprr(row[4].value),
                    reprr(data),
                    reprr(dataehora),
                    reprr(row[7].value),
                    reprr(row[8].value),
                    reprr(row[9].value),
                    integer(row[10].value),
                    reprr(row[11].value),
                    integer(row[12].value),
                    integer(row[13].value),
                    integer(row[14].value),
                    integer(row[15].value),
                    integer(row[16].value),
                    integer(row[17].value),
                    integer(row[18].value),
                    integer(row[19].value),
                    integer(row[20].value),
                    integer(row[21].value),
                    integer(row[22].value),
                    integer(row[23].value),
                    integer(row[24].value),
                    integer(row[25].value),
                    integer(row[26].value),
                    integer(row[27].value),
                    integer(row[28].value),
                    integer(row[29].value),
                    integer(row[30].value),
                    reprr(row[31].value),
                    reprr(row[32].value),
                    integer(row[33].value),
                    reprr(row[34].value),
                    reprr(row[35].value),
                    reprr(row[36].value),
                )
            except:
                print(row[36].value)
            print(q)
    if str(sheet.title)=="Vítimas":
        for row in sheet.rows:
            dias=str(row[10].value).split("/")
            if len(dias)==3:
                data="'%s-%s-%s'"%(dias[2],dias[1],dias[0],)
            else:
                data="NULL"
            q="INSERT INTO vitimas(id_acidente,id_vitima,id_veiculo,sexo,idade,tipo_vitima,classificacao,tipo_veiculo,estado_alcoolizacao,escolaridade,data) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s);"%(
                reprr(row[0].value),
                integer(row[1].value),
                integer(row[2].value),
                reprr(row[3].value),
                integer(row[4].value),
                reprr(row[5].value),
                reprr(row[6].value),
                reprr(row[7].value),
                reprr(row[8].value),
                integer(row[9].value),
                data,
            )
            print(q)


            #   wb2 = load_workbook('test.xlsx')
#print wb2.get_sheet_names()